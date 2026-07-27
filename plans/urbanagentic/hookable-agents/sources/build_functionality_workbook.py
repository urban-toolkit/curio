#!/usr/bin/env python3
"""Build the Agentic Functionality Workbook.

One data model (FLOWS) drives both:
  1. numbered circular callout overlays baked onto each screen, and
  2. the Excel workbook text,
so the numbers on the images always match the explanations.

Run:
    python3 sources/build_functionality_workbook.py
Outputs:
    workbook-assets/NN-*.annotated.png
    Agentic-Functionality-Workbook.xlsx
"""
from __future__ import annotations

from pathlib import Path

from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
PNG = ROOT / "png-concepts"
ASSETS = ROOT / "workbook-assets"
RUBIK = Path(__file__).resolve().parent / "fonts" / "Rubik.ttf"
OUT_XLSX = ROOT / "Agentic-Functionality-Workbook.xlsx"

# ── palette ───────────────────────────────────────────────────────────────
DARK = "1E1F23"
GREEN = "2F8F4A"
BLUE = "3567C7"
ORANGE = "E86A3C"
PURPLE = "7A4BD1"
INK = "1E1F23"
SUBTLE = "6B6B76"
LINE = "D0D0D5"
BAND = "F4F4F6"
WHITE = "FFFFFF"

ACCENT_RGB = {
    DARK: (30, 31, 35), GREEN: (47, 143, 74), BLUE: (53, 103, 199),
    ORANGE: (232, 106, 60), PURPLE: (122, 75, 209),
}

# ── flow data model (callouts drive both image overlays and workbook text) ──
FLOWS = [
    {
        "id": "01", "png": "01-agents-catalog-drawer.png", "accent": DARK,
        "title": "Agents Catalog — three-scope lifecycle drawer",
        "intro": "The right sidebar drawer separates three lifecycle scopes (Global Catalog · My Imports · Installed in this project) but uses the EXACT card action controls of the Data / Node catalogs: a dark 'Install' primary, a neutral white-outline 'Uninstall' secondary, the shared 'Publish' → 'Published' pill (My Imports), 'Delete', and an 'Import package' footer. Global Catalog is shown here.",
        "callouts": [
            (1210, 33, "Agents Catalog drawer", "The right drawer is the Agents Catalog, mirroring Curio's Data / Node catalog pattern."),
            (1300, 128, "Search + Sort", "Find an agent by name, hook, or keyword, and reorder the list."),
            (1300, 172, "Scope tabs", "Global Catalog · My Imports · Installed in this project — the three lifecycle scopes; Global is active here."),
            (1250, 204, "Category filters", "Scope the list by category: All, Data, Node, Canvas, Package, Evaluate."),
            (1600, 280, "Install / Uninstall", "Same controls as the Data / Node catalog cards: a dark 'Install' primary for available agents, a neutral white-outline 'Uninstall' for installed ones. No bespoke labels, colors, or per-card settings button."),
            (1140, 494, "Selected state", "The chosen card is outlined with a colored left bar (Connection Builder, available to install)."),
            (145, 300, "Node + agents palette", "Curio's dark left palette: built-in nodes plus DATA / PACKAGES / AGENTS dropdowns."),
            (1330, 905, "Import package", "Footer action, matching the Data / Node catalog's Import footer; validates into private My Imports only."),
        ],
        "states": [
            "Card actions reuse the catalog primitives: dark Install / Update, neutral Uninstall / Unpublish / Delete, and the shared Publish → Published pill.",
            "Installing scopes a ProjectAgentTemplate to the active project; it appears in that project's AGENTS palette.",
            "Global Catalog: Install / Uninstall (no user Publish). My Imports: Install + Publish/Published pill + Delete. Installed in this project: Uninstall.",
        ],
        "notes": [
            "Reuses the Data / Node Catalog action controls exactly — same styling, placement, pill states, and labels; no new or inconsistent action patterns.",
            "Global/built-in items are never user-republished; only owned My Imports definitions expose the Publish pill.",
        ],
    },
    {
        "id": "02", "png": "02-main-dataflow-attached-agents.png", "accent": DARK,
        "title": "Attached agents in a dataflow",
        "intro": "Attached agents float as macOS-Dock-style tiles — compact square, icon-only, left-aligned beneath their node (no background shelf). They magnify on hover and show the agent name in a tooltip. Node agents dock beneath their node; whole-dataflow agents dock on the canvas.",
        "callouts": [
            (270, 500, "Node agent dock", "An attached node agent floats as a square, icon-only tile left-aligned beneath its node (green Dataset Finder on Data Loading)."),
            (898, 206, "Canvas agent dock", "Whole-dataflow agents (Dataflow Builder, Validation, Optimization) float near the top of the canvas."),
            (827, 146, "Hover tooltip", "Hovering a tile magnifies it and reveals the agent name — the macOS Dock interaction pattern."),
            (827, 248, "Running indicator", "A small dot beneath a tile marks a running / active agent."),
            (475, 777, "Neighbour falloff", "Several agents on one node float as square tiles; the hovered one enlarges most and neighbours a little (Dock magnification)."),
            (1300, 274, "Catalog stays open", "The catalog remains available to attach more agents."),
            (700, 405, "Agents layer on the flow", "Tiles float over the dataflow without hiding it; a click opens the agent's chat session."),
        ],
        "states": [
            "Dock tiles persist after selection changes; hover magnifies the item under the pointer.",
            "The canvas dock represents agents hooked to the entire graph.",
        ],
        "notes": [
            "Attached-agent visibility is a core rule: ownership is always obvious.",
            "Tiles are visually identical; agent type is carried by the bot-head icon + accent. A running dot shows active state.",
        ],
    },
    {
        "id": "03", "png": "03-attachment-dataset-finder-to-data-load.png", "accent": GREEN,
        "title": "Attach Dataset Finder — two-lane suggestions",
        "intro": "Attaching Dataset Finder opens its chat. It does discovery + selection only, proposing two lanes in one card — External sources (which hand off to Node Builder) and datasets From your Data Catalog (which reuse the install flow). Nothing changes until you confirm.",
        "callouts": [
            (1275, 33, "Agent header", "One top header (DEC-042): ‹ › agent cycling, bot icon + name + position (1/4), and Close — no Pin here, and no static 'Agents Catalog' bar (that chrome is roster-drawer-only)."),
            (1262, 72, "Identification details", "The attached target and session chip form the header's second line — the content below is unchanged."),
            (1390, 139, "Initial intent", "The starting instruction for this attachment is pinned in an editable field directly below the title and cycling controls."),
            (1533, 88, "Session id", "Keyed by node id · agent type · project id, so each attachment is a distinct, resumable session."),
            (1174, 250, "Agent reply", "The agent responds to the intent in the transcript, like Cursor or Claude."),
            (1300, 330, "External lane → Node Builder", "External sources (API / Portal) with a confidence score; selecting one hands off to Node Builder to build the fetch node. No inline preview link."),
            (1300, 456, "Catalog lane → install", "Datasets already in your Data Catalog, each with a format badge and an install-state chip (installed / not installed); selecting reuses the install flow. No inline preview link."),
            (1300, 863, "Suggested prompts", "Alternatives as a 'SUGGESTED PROMPTS' chip row (e.g. 'Only build NOAA', 'Catalog only') — the card carries no action button."),
            (1390, 905, "Confirm from the input", "The primary confirm is a suggested prompt prefilled (editable) in the chat input — 'Build the NOAA node and install Heat Advisory Days' — reviewed and sent by the user; nothing changes silently."),
            (283, 500, "Attachment dock item", "The agent's floating square tile, left-aligned beneath the node; clicking it opened this session."),
        ],
        "states": [
            "Discovery yields two lanes: external sources (→ Node Builder) and Data Catalog datasets (→ install).",
            "Per-row badges: source-type + confidence (external); format + install-state chip (catalog).",
            "No card action button — the confirm is a suggested, editable prompt in the chat input.",
        ],
        "notes": [
            "The Dataset Finder never authors fetch code — that is Node Builder's job (see memo 00-dataset-finder-workflow-replan).",
            "Suggestion rows have no inline preview link — dataset preview / detail is reviewed later via the Data Catalog drawer detail modal / screen, so the entry point isn't duplicated.",
            "Generic pattern: confirmations / next steps are suggested prompts in the input, not bespoke buttons — shared by every agent.",
        ],
    },
    {
        "id": "04", "png": "04-attachment-node-explainer-to-node.png", "accent": BLUE,
        "title": "Attach Node Explainer — drag from palette",
        "intro": "Attachment is performed by dragging an installed agent from the AGENTS palette onto a node — exactly like adding a node or dataset. Here Node Explainer is dragged onto a compute node.",
        "callouts": [
            (435, 200, "AGENTS palette", "The drag starts from the AGENTS palette (installed agents only)."),
            (500, 380, "Drag path", "A dotted path follows the drag from the palette toward the target."),
            (573, 556, "Dragged agent", "The picked-up agent tile (Node Explainer) being dropped onto the node."),
            (573, 648, "Compatible node", "The compute node highlights (blue) as a valid drop target."),
            (427, 494, "Drag to attach", "A callout confirms attachment is by dragging from the palette."),
        ],
        "states": [
            "Attachment is a palette drag; compatible targets highlight during the drag.",
            "The agent must be installed (in the palette) before it can be dragged.",
        ],
        "notes": [
            "Consistent with the datasets/nodes model: install in the catalog, drag from the palette.",
            "Blue is reserved for inspection / explanation agents.",
        ],
    },
    {
        "id": "05", "png": "05-attachment-dataflow-builder-to-canvas.png", "accent": ORANGE,
        "title": "Attach Dataflow Builder — drag to canvas",
        "intro": "A canvas-level agent is attached the same way — by dragging an installed agent from the AGENTS palette onto the canvas. Here the Dataflow Builder is dragged onto the whole-canvas hook.",
        "callouts": [
            (435, 200, "AGENTS palette", "The drag starts from the AGENTS palette (installed agents only)."),
            (540, 340, "Drag path", "A dotted path follows the drag toward the canvas."),
            (655, 435, "Dragged agent", "The picked-up Dataflow Builder tile being dropped onto the canvas."),
            (245, 160, "Canvas hook boundary", "A dashed orange boundary shows the whole-canvas drop target."),
            (436, 278, "Drag to attach", "A callout confirms attachment is by dragging from the palette."),
        ],
        "states": [
            "Canvas agents attach by dragging onto the canvas hook boundary.",
            "The agent must be installed (in the palette) before it can be dragged.",
        ],
        "notes": [
            "Consistent with the datasets/nodes model: install in the catalog, drag from the palette.",
            "Orange is Curio's primary action / canvas-level accent.",
        ],
    },
    {
        "id": "06", "png": "06-agent-refinement-sidebar-open.png", "accent": BLUE,
        "title": "Unified agent chat — Node Explainer",
        "intro": "The same chat drawer, a different agent. Node Explainer explains the node and lets you configure behavior — all in chat. It is visually indistinguishable from Dataset Finder; only the agent type, config, and history differ.",
        "callouts": [
            (1275, 33, "Agent header", "The same single top header as sheet 03 (DEC-042): identity + cycling + Close live in the header; only the agent changes."),
            (1145, 33, "Cycle agents", "Prev / next moves through all attached agents; this is agent 2 of 4."),
            (1262, 72, "Identification details", "Attached target + session chip on the header's second line (Node Explainer, 2/4) — same layout as sheet 03."),
            (1390, 139, "Initial intent", "The starting instruction is pinned below the title / cycling controls — same field, different agent."),
            (1533, 88, "Session id", "node id · agent type · project id — here node 4 · Node Explainer."),
            (1390, 211, "What it can read", "A system line states context and policy (reads code / output / lineage; suggestions only)."),
            (1174, 255, "Explanation in chat", "The agent answers in the transcript instead of a bespoke panel."),
            (1390, 315, "Behavior config in chat", "Behavior is configured with in-chat quick replies (Planner-friendly selected)."),
            (1290, 863, "Quick replies", "One-tap refinements like 'Use technical style' or 'List assumptions'."),
            (1390, 905, "Chat input", "Everything — refine, reconfigure, re-run — happens through chat."),
            (463, 775, "Open dock item", "The magnified dock tile marks the open session and links to the chat drawer."),
        ],
        "states": [
            "Clicking the tab opens this agent's session; the header arrows switch agents.",
            "Behavior and prompt are set by chatting, not by a separate form.",
            "The transcript is the run history for what this agent did on this node.",
        ],
        "notes": [
            "Node Explainer and Dataset Finder share one drawer — they differ only by type, config, and history.",
            "Review policy (suggestions only) is stated in-line; no destructive edits.",
        ],
    },
    {
        "id": "07", "png": "07-dataset-finder-overview.png", "accent": GREEN,
        "title": "Agent session — both selection paths",
        "intro": "Each attachment is a persistent session; the transcript IS the run history. This one shows both paths in one conversation: an external pick hands off to Node Builder, and a Data Catalog pick is auto-installed via the existing install flow.",
        "callouts": [
            (1275, 33, "Agent header", "The opened agent view's single top header (DEC-042): identity, cycling, session details, and Close — no Pin, no static 'Agents Catalog' bar."),
            (1262, 72, "Identification details", "Attached target + session chip on the header's second line; the transcript below is unchanged."),
            (1390, 139, "Initial intent", "The session's starting intent stays pinned below the title while the conversation continues."),
            (1533, 88, "Session id", "The tuple uniquely identifies this attachment; the transcript is its history."),
            (1390, 300, "Two-lane suggestions", "The earlier suggestion turn, kept in the transcript — external sources and Data Catalog matches."),
            (1578, 588, "Mixed selection", "The user picks one external ('build the NOAA node') and one catalog dataset ('install Census ACS')."),
            (1300, 705, "Hand-off to Node Builder", "The external pick spawns a Node Builder session (installed in project, reviewed) to build the fetch node — the Finder does not write code."),
            (1300, 815, "Catalog auto-install", "The catalog pick reuses the existing dataset install flow (project-scoped) and attaches — no code generated."),
            (1390, 905, "Continue the session", "Keep refining in the same chat; nothing is a separate page."),
        ],
        "states": [
            "One session, two outcomes: external → Node Builder hand-off; catalog → auto-install.",
            "Re-opening the tab resumes the same session with its full history.",
            "The Finder delegates implementation and installation; it never authors fetch code.",
        ],
        "notes": [
            "Run history = chat history; there is no separate history panel.",
            "The hand-off card and install card make the two routes explicit in the transcript.",
        ],
    },
    {
        "id": "08", "png": "08-dataset-review-modal.png", "accent": BLUE,
        "title": "Node Builder — generated dataset node",
        "intro": "When an external source is selected, the Dataset Finder hands off to Node Builder, which generates the complete dataset node. Its own chat session posts the node inline — request code, API-key field, request params, a parsing / error-handling / output-format checklist, and a data sample — for review before it is added.",
        "callouts": [
            (1275, 33, "Agent header", "The single top header (DEC-042) now shows this session's agent identity + Close; no Pin, no static 'Agents Catalog' bar."),
            (1262, 87, "Node Builder session", "A distinct session (node · node-builder · project) spawned by Dataset Finder for the selected source."),
            (1390, 211, "Spawned + review policy", "A system line notes it was spawned by Dataset Finder and runs review-before-apply."),
            (1276, 330, "Generated node preview", "The exact Data Loading node Node Builder will create — not authored by the Finder."),
            (1350, 355, "Request code + params", "The generated request code and the request parameters (datasetid · stationid · dates · units)."),
            (1248, 427, "API key input", "Auth is surfaced inline as a required field; stored per dataflow, never written to the workflow file."),
            (1300, 478, "Parse / error / output", "A checklist confirms response parsing, error handling, and output formatting are included."),
            (1390, 520, "Data sample", "A few preview rows so you know what you'll get. The card has no action button."),
            (1390, 905, "Add via suggested prompt", "Adding is a suggested prompt prefilled (editable) in the chat input — 'Add the NOAA dataset node to Data Loading' — with a 'Not now' chip alternative; there is no card button."),
        ],
        "states": [
            "External-source implementation is Node Builder's output, not the Finder's.",
            "The node is fully reviewable — code, params, auth, parsing, error handling, output — before adding.",
            "Adding is a suggested, editable prompt in the input; missing credentials block the send until provided.",
        ],
        "notes": [
            "Screen 08 is now a Node Builder session; the Dataset Finder only selected the source.",
            "No card action button — the add / dismiss actions are suggested prompts (generic pattern).",
        ],
    },
    {
        "id": "09", "png": "09-datasets-created-in-palette.png", "accent": GREEN,
        "title": "Datasets created & provenance (two origins)",
        "intro": "After you confirm, both results land in Curio's real DATA palette — but from two origins: the external source is a Node Builder-authored node (EXTERNAL · Node Builder), and the catalog dataset was auto-installed via the install flow (IMPORTED · Data Catalog). A run-history entry is logged for each.",
        "callouts": [
            (318, 105, "Confirmation toast", "Confirms the NOAA node was created (Node Builder) and Census ACS was installed from the catalog."),
            (271, 150, "DATA palette (open)", "The DATA trigger opens (peach, active, chevron up) to reveal the installed datasets."),
            (258, 300, "Datasets dropdown", "Matches Curio's real dark DATA dropdown: 'DATASETS' header, installed count, dataset rows, and a Browse Data Catalog action."),
            (516, 366, "External node (Node Builder)", "The NOAA row carries a blue accent, an EXTERNAL status pill, and a 'Node Builder' chip — the agent-authored fetch node."),
            (516, 452, "Catalog install (Data Catalog)", "The Census ACS row carries an orange accent, an IMPORTED status pill, and a 'Data Catalog' chip — auto-installed, not generated."),
            (600, 548, "Computed dataset", "Existing computed datasets show a COMPUTED pill + usage counts and no agent chip, so origin stays legible."),
            (300, 872, "Browse Data Catalog", "Peach footer action to open the full Data Catalog."),
            (1405, 266, "Both logged in the session", "A result turn (node created) and an install turn (from the catalog) are logged in the chat — the transcript is the run history."),
        ],
        "states": [
            "External source → Node Builder-authored node (EXTERNAL); catalog dataset → auto-install (IMPORTED).",
            "The dark dropdown matches the live Curio DATA palette layout.",
            "The two origins are visually distinguished from each other and from computed datasets.",
        ],
        "notes": [
            "Dropdown layout aligned to png-ideas/datasets_palette_dropdown (source of truth).",
            "Provenance is explicit: Node Builder authorship (EXTERNAL) vs Data Catalog install (IMPORTED).",
        ],
    },
    {
        "id": "10", "png": "10-dataflow-builder-orchestration.png", "accent": ORANGE,
        "title": "Dataflow Builder — orchestration",
        "intro": "The Dataflow Builder is the master orchestration agent. In the same unified chat drawer it interprets the objective, plans subtasks, proposes a reviewed Install in project for any missing specialist (never auto-installing), coordinates the specialists installed in this project (with live status), then merges, evaluates, and delivers an executable dataflow.",
        "callouts": [
            (1275, 33, "Agent header", "The orchestrator opens in the unified chat drawer like any agent — the single top header (DEC-042) shows its identity + Close."),
            (1262, 87, "Master orchestrator", "The Dataflow Builder coordinates other agents instead of recommending datasets / nodes / packages itself."),
            (1390, 139, "Objective / intent", "The user's high-level outcome, pinned as the initial intent."),
            (1533, 88, "Session id", "Canvas-scoped session: node (canvas) · agent type · project id."),
            (1300, 288, "Execution plan", "The objective is decomposed into ordered subtasks (a reusable Planning capability)."),
            (1390, 477, "Reviewed Install in project", "For any specialist the project is missing, it proposes a reviewed Install in project — never an automatic or silent install; installed specialists appear in the project palette."),
            (1300, 552, "Specialized agents", "It coordinates the specialists installed in this project (Dataset Finder, Node Builder, Connection Builder, Validation, Node Explainer)."),
            (1600, 580, "Live status", "Per-agent status — done / running / queued — including persistent background execution."),
            (1174, 767, "Merge & evaluate", "It merges outputs and evaluates coherence, refining before changing the graph (review-before-apply)."),
            (1300, 863, "Orchestration controls", "Quick replies: show the plan, pause, or run in the background."),
            (827, 205, "Canvas dock", "The orchestrator is a canvas-level agent; its dock tile links to this chat."),
        ],
        "states": [
            "Objective → plan → reviewed Install in project for missing specialists → coordinate → merge → evaluate → refine → deliver.",
            "Missing specialists get a reviewed Install in project (never automatic); installed specialists run (often in the background) with visible status; graph changes are confirmed.",
            "Agent implementation (LangChain) is separate from this assignment / UI — see docs 09 & 10.",
        ],
        "notes": [
            "Reuses the existing chat drawer, dock, and card components — no new layout.",
            "Redefines the Dataflow Builder from a recommender to the master orchestrator (memo 00-dataflow-builder-orchestration).",
        ],
    },
    {
        "id": "11", "png": "11-agents-palette.png", "accent": ORANGE,
        "title": "Agents palette — installed agents",
        "intro": "The AGENTS palette lists ONLY the agents installed in this project, using the dataset palette row exactly: an icon tile, the agent name, an agent.<id>@<v> reference, and a category chip — no per-row publish/install action, no accent strip, and no drag-handle icon (publish lives in the drawer). Global agents are installed from the right-sidebar Agents Catalog (screen 01) first; then they appear here and can be dragged onto the dataflow to attach.",
        "callouts": [
            (435, 165, "AGENTS palette", "Agents get their own palette dropdown beside DATA and PACKAGES — the same access pattern as nodes and datasets."),
            (500, 314, "Installed only", "The palette lists only agents installed in this project — no global-catalog agents (those live in the drawer)."),
            (620, 385, "Agent reference", "Each row shows an agent.<id>@<v> reference, like the computed.<id>@<v> reference on a dataset row."),
            (497, 402, "Category chip", "A category chip (Canvas / Data / Node / Evaluate) — the analog of the dataset provenance chip."),
            (540, 360, "Whole row draggable", "The row has no drag-handle icon, no accent strip, and no publish/install action — it is dragged as a whole onto a node or the canvas, exactly like a dataset row."),
            (616, 780, "Drag hint", "Reminder that installed agents are attached by dragging them onto the dataflow."),
            (616, 872, "Browse Agents Catalog", "Opens the right-sidebar Agents Catalog to install more global agents."),
            (557, 105, "Install first, then drag", "Install global agents from the catalog; drag installed ones to attach; refine in the right drawer."),
        ],
        "states": [
            "Palette access: an AGENTS dropdown beside DATA and PACKAGES lists installed agents only.",
            "Rows mirror the dataset palette exactly: icon + name + reference + category chip — no accent strip, no drag-handle icon, no publish/install action.",
            "Install in the catalog → appears in the palette → drag to attach; refinement continues in the right drawer.",
        ],
        "notes": [
            "Catalog installs; the palette holds installed agents — exactly the datasets/nodes model.",
            "Publish lives solely in the drawer (like datasets/node packages); no palette publish control and no scope tiers.",
        ],
    },
    {
        "id": "12", "png": "12-agent-settings-cost.png", "accent": GREEN,
        "title": "Agent settings — Cost",
        "intro": "The shared six-screen agent settings modal, opened from the project-installed detail, over the current surface. Cost shows per-run and rolling budgets, current usage, alert thresholds, the pricing effective date, and explicit Estimated vs Actual labels.",
        "callouts": [
            (360, 152, "Modal over surface", "Agent governance opens as a modal over the current canvas, not by replacing the drawer."),
            (430, 214, "Six-screen rail", "A left rail lists the six settings screens (Cost active); the same shell is reused everywhere."),
            (760, 206, "Budget & spend", "Per-run and rolling (30-day) budgets, with the pricing effective date."),
            (970, 344, "Usage meter", "Current spend against the rolling budget, with a percent-used bar."),
            (700, 440, "Estimated vs Actual", "Explicit labels separate estimates from provider-reported actuals."),
            (430, 782, "Scope ceilings", "Account policy sets ceilings; this scope can only tighten them."),
        ],
        "states": ["Settings are a modal shell, not a separate page.", "Cost defaults are materialized per scope, bounded by account ceilings."],
        "notes": ["One shared modal across agents; per-agent defaults, never a generic global preset.", "Estimate vs actual is always labelled."],
    },
    {
        "id": "13", "png": "13-agent-settings-quotas.png", "accent": GREEN,
        "title": "Agent settings — Quotas",
        "intro": "Quota limits for executions, tokens, tool calls, concurrency, and rate window, with reservations and a reset window. Values inherited from account policy are shown locked.",
        "callouts": [
            (430, 214, "Six-screen rail", "Quotas active in the shared settings shell."),
            (760, 214, "Limits", "Executions, tokens, tool calls, concurrency, and rate window."),
            (1240, 250, "Inherited / locked", "Limits inherited from account policy are shown with a lock chip."),
            (760, 480, "Reservations", "Capacity held back for retries."),
            (1300, 206, "Reset window", "When the quota window resets."),
        ],
        "states": ["Quotas can tighten but not exceed account ceilings.", "Locked values are inherited, not editable at this scope."],
        "notes": ["Reset provenance and reservations are shown.", "Same shell chrome as every settings screen."],
    },
    {
        "id": "14", "png": "14-agent-settings-resource-policies.png", "accent": GREEN,
        "title": "Agent settings — Resource policies",
        "intro": "Provider profile, model, and processing locality (Local vs Remote), plus context/output, time, egress, and tools/network bounds. Providers are authorized profiles, not key fields; secret values are never shown.",
        "callouts": [
            (760, 206, "Provider profile", "An authorized profile identifies provider/model — not raw key fields."),
            (980, 340, "Local vs Remote", "Processing locality toggle; local never silently falls back to remote."),
            (760, 420, "Bounds", "Context/output, time limit, egress, and tools/network restrictions."),
            (1300, 206, "Secrets never shown", "Secret values are never returned or rendered."),
        ],
        "states": ["Locality is explicit; failures never silently fall back.", "No secret is ever displayed or stored in the attachment."],
        "notes": ["Providers are profiles, not key fields.", "Bounds tighten within account ceilings."],
    },
    {
        "id": "15", "png": "15-agent-settings-prompt-quality.png", "accent": GREEN,
        "title": "Agent settings — Prompt quality",
        "intro": "A pinned validation suite, rubric, and pass threshold, with recent evaluation results (passed / running), findings, and evaluation usage/cost.",
        "callouts": [
            (760, 214, "Pinned suite", "Validation suite + rubric + pass threshold."),
            (760, 360, "Evaluations", "Recent eval results with pass / running status and scores."),
            (760, 520, "Eval usage", "Token and cost usage of evaluation runs."),
        ],
        "states": ["Evaluations have queued/running/completed/failed states.", "Findings and usage are visible for the pinned suite."],
        "notes": ["Prompt quality is owned by the imported definition; project/attachment views are read-only.", "Distinct from the audit record."],
    },
    {
        "id": "16", "png": "16-agent-settings-prompt-editor.png", "accent": GREEN,
        "title": "Agent settings — Prompt editor",
        "intro": "An owned My-Imports prompt draft editor with variables/schema validation and a diff. Saving a draft never mutates the immutable imported definition artifact.",
        "callouts": [
            (760, 214, "Draft editor", "Owner-authorized prompt draft with variables and schema checks."),
            (760, 380, "Validation + diff", "Variables valid, schema valid, and a diff summary."),
            (760, 420, "Immutable definition", "Saving a draft never mutates the immutable imported definition bytes."),
            (1300, 405, "Save draft", "A generic system control (permitted in the settings shell)."),
        ],
        "states": ["Only an owned imported definition exposes the editor.", "Draft save is separate from release; bytes stay immutable."],
        "notes": ["Prompt authoring belongs to the owned import, not project/attachment scopes.", "Save/Cancel/Save draft are allowed system controls."],
    },
    {
        "id": "17", "png": "17-agent-settings-prompt-audit.png", "accent": GREEN,
        "title": "Agent settings — Prompt audit",
        "intro": "Versioned privacy / security / compliance rules with audit runs and findings, plus an append-only governance history that is distinct from the chat/execution transcript.",
        "callouts": [
            (760, 214, "Versioned rules", "Privacy, security, and compliance rules with per-version findings."),
            (760, 420, "Audit history", "Append-only, authorized governance history."),
            (760, 520, "Distinct from transcript", "Prompt Audit is separate from the execution/chat transcript."),
        ],
        "states": ["Audit is access-controlled and append-only.", "Findings are tracked per rule version."],
        "notes": ["Governance history is not the run history.", "Read-only outside the owning import."],
    },
    {
        "id": "18", "png": "18-agents-drawer-lifecycle.png", "accent": ORANGE,
        "title": "Agents drawer — lifecycle",
        "intro": "The agent lifecycle as a vertical flow across the three scopes: Import package creates a private My-Imports definition; Publish promotes it to the Global Catalog; Install in project creates a ProjectAgentTemplate; Attach creates a private instance. Each command is explicit and independent.",
        "callouts": [
            (1300, 141, "My Imports · private", "Import package creates a private account definition — it never installs or publishes."),
            (1470, 248, "Publish", "Publish promotes a private import to the Global Catalog."),
            (1300, 288, "Global Catalog", "Built-in + published definitions, reusable across projects."),
            (1450, 393, "Install in project", "Install creates a ProjectAgentTemplate scoped to this project."),
            (1300, 433, "Installed in this project", "Shows in this project's palette; Project agent settings + Uninstall from project."),
            (1300, 580, "Attached instance", "Dragging to a target creates a private instance — no Publish / Share."),
        ],
        "states": ["Import, Publish, Install in project, Attach are independent explicit commands.", "Installation is project-scoped; attachments alone are dataflow-scoped."],
        "notes": ["Reconciled project-scoped model (docs 03 / 11).", "Import lands only in private My Imports; neither triggers the other."],
    },
    {
        "id": "19", "png": "19-settings-scope-applicability.png", "accent": DARK,
        "title": "Settings — scope applicability",
        "intro": "One matrix showing how the six settings screens apply across four ownership scopes: Account policy owns ceilings; the imported definition owns the prompt screens; the project agent owns project defaults; the attached instance can only tighten and reads prompt screens as evidence.",
        "callouts": [
            (1300, 162, "Applicability matrix", "Which settings each ownership scope edits, inherits, or reads."),
            (625, 224, "Account policy", "Owns ceilings for Cost / Quotas / Resources."),
            (835, 224, "Imported definition", "Owns the prompt screens (quality / editor / audit)."),
            (1043, 224, "Project agent", "Owns project Cost/Quotas/Resource defaults; prompt screens read-only."),
            (1252, 224, "Attached instance", "Can only tighten; prompt screens are read-only evidence."),
        ],
        "states": ["Account ceilings bound everything below.", "Prompt screens are editable only in their owning import."],
        "notes": ["Attachment policy tightens; it never loosens.", "Read-only cells are evidence, not editable."],
    },
    {
        "id": "20", "png": "20-node-explainer-only-workflow.png", "accent": BLUE,
        "title": "Node Explainer — no node tab",
        "intro": "Nodes have no built-in Explanation tab, configuration, direct LLM call, or explanation cache. 'Explain with Node Explainer' opens the normal project install, attach, and unified-chat workflow — not a bespoke node panel.",
        "callouts": [
            (685, 573, "No Explanation tab", "Node cards carry no built-in Explanation tab."),
            (685, 606, "Explain with Node Explainer", "Opens the normal install, attach, and chat workflow."),
            (435, 285, "Agent, not a tab", "Explanation is an attached agent, refined in the unified chat."),
        ],
        "states": ["No node-level explanation state or cache exists.", "Canvas/full-flow explanation is the separate Dataflow Explainer behavior."],
        "notes": ["The removed node tab is never re-presented.", "Explanation follows the standard agent lifecycle."],
    },
    {
        "id": "21", "png": "21-agents-catalog-my-imports-publish.png", "accent": BLUE,
        "title": "Agents Catalog — My Imports · publish to the Catalog Hub",
        "intro": "The My Imports scope of the Agents Catalog drawer holds the account's own imported private definitions and exposes the SAME catalog publishing controls datasets and node packs use: a dark 'Install' primary, the shared 'Publish' → 'Published' pill that lists the definition in the global Catalog Hub, and a neutral 'Delete'. Publishing is imported-only and lives here — never on Global Catalog or Installed cards.",
        "callouts": [
            (1300, 172, "My Imports scope", "The account's own imported private definitions — the only place the Publish control appears."),
            (1600, 272, "Install", "Install this owned definition as a project template, the same dark primary used across the catalogs."),
            (1600, 298, "Publish → Catalog Hub", "The shared CatalogPublishPill: 'Publish' (sky) lists the owned definition in the global Catalog Hub so other users can discover and install it, exactly like datasets and node packs."),
            (1600, 391, "Published state", "Once listed, the pill becomes a neutral 'Published' badge — the same published-state pill as the dataset / node-pack catalogs (Dataset Finder and Validation shown published)."),
            (1600, 324, "Delete", "Remove the owned import; a distinct, authorized operation separate from Uninstall and Unpublish."),
            (1430, 271, "Private definition", "A 'Private' pill marks an account-scoped import; publishing is what makes it discoverable in the Hub."),
            (1250, 240, "install · publish · delete", "The My Imports action set, matching the established catalog drawer action patterns."),
        ],
        "states": [
            "My Imports cards expose Install + Publish/Published pill + Delete — the same publishing controls and states as the Data / Node catalogs.",
            "Publish lists an owned validated imported definition in the global Catalog Hub; Published shows it is already listed.",
            "Only owned My Imports definitions are publishable; Global Catalog and Installed cards never show Publish.",
        ],
        "notes": [
            "Restores the imported-agent publishing flow using the exact CatalogPublishPill (Publish → Published) datasets and node packs use — no bespoke control.",
            "The publishing action is never removed or hidden for imported agents; it is the imported-only path to the global Catalog Hub.",
        ],
    },
]

# ── numbered circular callout overlays ──────────────────────────────────────
def _badge_font(px: int) -> ImageFont.FreeTypeFont:
    try:
        f = ImageFont.truetype(str(RUBIK), px)
        try:
            f.set_variation_by_name("Bold")
        except Exception:
            pass
        return f
    except OSError:
        return ImageFont.load_default()


def draw_badge(img: Image.Image, x: int, y: int, n: int, rgb: tuple[int, int, int]) -> None:
    """Supersampled numbered circle: soft shadow, white ring, colored fill."""
    S = 4
    d = 38
    tile = Image.new("RGBA", (d * S, d * S), (0, 0, 0, 0))
    t = ImageDraw.Draw(tile)
    m = 3 * S  # margin for shadow/ring
    t.ellipse((m + 2 * S, m + 3 * S, d * S - m + 2 * S, d * S - m + 3 * S), fill=(0, 0, 0, 70))
    t.ellipse((m, m, d * S - m, d * S - m), fill=(255, 255, 255, 255))
    r = 2 * S
    t.ellipse((m + r, m + r, d * S - m - r, d * S - m - r), fill=rgb + (255,))
    f = _badge_font(17 * S)
    t.text((d * S / 2, d * S / 2 - S), str(n), font=f, fill=(255, 255, 255, 255), anchor="mm")
    tile = tile.resize((d, d), Image.Resampling.LANCZOS)
    img.paste(tile, (int(x - d / 2), int(y - d / 2)), tile)


def annotate() -> dict[str, Path]:
    ASSETS.mkdir(parents=True, exist_ok=True)
    out = {}
    for flow in FLOWS:
        img = Image.open(PNG / flow["png"]).convert("RGB")
        rgb = ACCENT_RGB[flow["accent"]]
        for i, (x, y, *_rest) in enumerate(flow["callouts"], start=1):
            draw_badge(img, x, y, i, rgb)
        dst = ASSETS / f"{flow['id']}.annotated.png"
        img.save(dst, optimize=True)
        out[flow["id"]] = dst
    return out


# ── workbook styling helpers ────────────────────────────────────────────────
def fill(hex_):
    return PatternFill("solid", fgColor=hex_)


def font(size=11, bold=False, color=INK):
    return Font(name="Rubik", size=size, bold=bold, color=color)


THIN = Side(style="thin", color=LINE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_MID = Alignment(wrap_text=True, vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")


def est_height(text: str, chars_per_line: int, base: int = 15, pad: int = 6, minimum: int = 20) -> int:
    lines = 0
    for para in str(text).split("\n"):
        lines += max(1, -(-len(para) // chars_per_line))
    return max(minimum, lines * base + pad)


def section_header(ws, row, label, accent):
    c = ws.cell(row=row, column=2, value=label)
    c.font = font(11, True, WHITE)
    c.fill = fill(accent)
    c.alignment = Alignment(vertical="center", indent=1)
    for col in (3, 4):
        ws.cell(row=row, column=col).fill = fill(accent)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 22


def bullets(ws, row, items, accent):
    for it in items:
        cell = ws.cell(row=row, column=2, value="•  " + it)
        cell.font = font(10, color="3A3A42")
        cell.alignment = WRAP
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.row_dimensions[row].height = est_height(it, 150)
        row += 1
    return row


# ── build ────────────────────────────────────────────────────────────────
def setup_flow_columns(ws):
    ws.column_dimensions["A"].width = 2.5
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 92
    ws.sheet_view.showGridLines = False


def build():
    assets = annotate()
    wb = Workbook()

    # ---- Overview -------------------------------------------------------
    ov = wb.active
    ov.title = "Overview"
    ov.sheet_properties.tabColor = DARK
    ov.sheet_view.showGridLines = False
    ov.column_dimensions["A"].width = 2.5
    ov.column_dimensions["B"].width = 26
    ov.column_dimensions["C"].width = 30
    ov.column_dimensions["D"].width = 74
    ov.merge_cells("B2:D2")
    t = ov["B2"]
    t.value = "Agentic Dataflows — Functionality Workbook"
    t.font = font(20, True, WHITE)
    t.fill = fill(DARK)
    t.alignment = Alignment(vertical="center", indent=1)
    ov.row_dimensions[2].height = 40
    ov.merge_cells("B3:D3")
    s = ov["B3"]
    s.value = "How reusable agents attach to Curio dataflows — discover, attach, configure, review, confirm, and track run history."
    s.font = font(11, color=SUBTLE)
    s.alignment = WRAP
    ov.row_dimensions[3].height = 30

    r = 5
    section_header(ov, r, "How to read this workbook", DARK); r += 1
    for it in [
        "Each numbered flow has its own tab (01–09), plus reference tabs for the agent catalog and the refinements made.",
        "Every screen carries numbered circular callouts. The numbers match the 'What it does' table on the same tab.",
        "Refinement is unified: every attached agent opens the SAME chat drawer (sheets 03, 06, 07, 08, 09). Attachments are visually indistinguishable and differ only by agent type, configuration, and chat history. Each is a session keyed by node id + agent type + project id, with prev/next navigation in the header.",
        "Callout color follows the agent: dark = catalog/overview, green = Dataset Finder, blue = Node Explainer, orange = Dataflow Builder, purple = Evaluator.",
        "Screens are design concepts rendered from the Curio UI baseline — not production screenshots.",
    ]:
        r = bullets(ov, r, [it], DARK)
    r += 1

    section_header(ov, r, "Color & agent legend", DARK); r += 1
    hdr = ["Accent", "Agent / use", "Meaning"]
    for i, h in enumerate(hdr):
        c = ov.cell(row=r, column=2 + i, value=h)
        c.font = font(10, True, WHITE); c.fill = fill(SUBTLE); c.border = BOX
    r += 1
    legend = [
        (GREEN, "Dataset Finder", "Data-oriented agents (sourcing datasets)."),
        (BLUE, "Node Explainer", "Inspection & explanation agents."),
        (ORANGE, "Dataflow Builder", "Canvas-level planning; primary action accent."),
        (PURPLE, "Validation / Package", "Evaluation, validation, and package agents."),
        (DARK, "Catalog / primary", "Neutral chrome and primary buttons."),
    ]
    for hexc, name, mean in legend:
        ov.cell(row=r, column=2).fill = fill(hexc)
        ov.cell(row=r, column=2).border = BOX
        cn = ov.cell(row=r, column=3, value=name); cn.font = font(10, True); cn.border = BOX; cn.alignment = WRAP_MID
        cm = ov.cell(row=r, column=4, value=mean); cm.font = font(10); cm.border = BOX; cm.alignment = WRAP
        ov.row_dimensions[r].height = 20
        r += 1
    r += 1

    section_header(ov, r, "Product thesis", DARK); r += 1
    for it in [
        "Urban Agentic is an agentic dataflow builder: give Curio an urban mission and it helps plan, revise, solve, evaluate, and explain an editable dataflow.",
        "Reusable agents are the building blocks. They are not isolated chatbots — they attach to specific Curio targets (datasets, nodes, canvas, connections, code, outputs, sources).",
        "Mission -> reusable agents -> hookable targets -> editable dataflow -> evaluation & explanation loop.",
    ]:
        r = bullets(ov, r, [it], DARK)
    r += 1

    section_header(ov, r, "Contents", DARK); r += 1
    for f in FLOWS:
        cn = ov.cell(row=r, column=2, value=f"{f['id']} · {f['title'].split(' — ')[0].split(' →')[0]}")
        cn.font = font(10, True); cn.alignment = WRAP_MID
        cd = ov.cell(row=r, column=3, value=f["title"]); cd.font = font(10, color=SUBTLE); cd.alignment = WRAP_MID
        ov.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        ov.row_dimensions[r].height = 18
        r += 1
    ov.freeze_panes = "A5"

    # ---- Agent Catalog & Hooks -----------------------------------------
    ac = wb.create_sheet("Agent Catalog & Hooks")
    ac.sheet_properties.tabColor = DARK
    ac.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", (2.5, 22, 12, 26, 26, 22)):
        ac.column_dimensions[col].width = w
    ac.merge_cells("B2:F2")
    h = ac["B2"]; h.value = "Reusable agents & hook targets"; h.font = font(16, True, WHITE); h.fill = fill(DARK)
    h.alignment = Alignment(vertical="center", indent=1); ac.row_dimensions[2].height = 32

    r = 4
    section_header2 = lambda row, label: (
        ac.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6),
        setattr(ac.cell(row=row, column=2), "value", label),
        setattr(ac.cell(row=row, column=2), "font", font(11, True, WHITE)),
        setattr(ac.cell(row=row, column=2), "fill", fill(DARK)),
        setattr(ac.cell(row=row, column=2), "alignment", Alignment(vertical="center", indent=1)),
    )
    section_header2(r, "Core agents"); ac.row_dimensions[r].height = 22; r += 1
    cols = ["Agent", "Category", "Primary hook", "Reads", "Review policy"]
    for i, cn in enumerate(cols):
        c = ac.cell(row=r, column=2 + i, value=cn); c.font = font(10, True, WHITE); c.fill = fill(SUBTLE); c.border = BOX; c.alignment = WRAP_MID
    r += 1
    agents = [
        ("Dataflow Builder", "Canvas", "Canvas (whole dataflow)", "mission, graph, run state; orchestrates specialized agents", "MASTER ORCHESTRATOR · confirm plan changes"),
        ("Dataset Finder", "Data", "Data Loading node", "mission, node context, catalog, geography, lineage", "Confirm before adding / replacing datasets"),
        ("Node Builder", "Node", "Canvas or selected connection", "schema, mission", "Review before creating node"),
        ("Connection Builder", "Node", "Canvas or selected nodes", "node types, data contracts", "Suggest / create valid connections"),
        ("Package Recommendation", "Package", "Node or canvas", "task, node context", "Recommend packages"),
        ("Validation", "Evaluate", "Node or full canvas", "code, coherence, data types, outputs, assumptions", "Report only"),
        ("Optimization", "Evaluate", "Canvas", "graph, performance, structure", "Suggest; confirm changes"),
        ("Node Explainer", "Node", "Any node w/ code, output, provenance", "node code, inputs, outputs, errors, lineage", "Suggestions only; no edits"),
    ]
    for row_vals in agents:
        for i, v in enumerate(row_vals):
            c = ac.cell(row=r, column=2 + i, value=v)
            c.font = font(10, bold=(i == 0)); c.border = BOX; c.alignment = WRAP
        ac.row_dimensions[r].height = est_height(max(row_vals, key=len), 30)
        r += 1
    r += 1
    section_header2(r, "Hook targets"); ac.row_dimensions[r].height = 22; r += 1
    for i, cn in enumerate(["Hook target", "Examples", "Context exposed"]):
        c = ac.cell(row=r, column=2 + i, value=cn); c.font = font(10, True, WHITE); c.fill = fill(SUBTLE); c.border = BOX
        c.alignment = WRAP_MID
    ac.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
    r += 1
    hooks = [
        ("Data Loading node", "CSV / DB / API fetch", "source, schema, preview, region, upstream lineage"),
        ("Analysis node", "Python, JS, transform, stats", "code, inputs, outputs, errors, lineage"),
        ("Visualization node", "map, chart, dashboard, HTML", "data bindings, encoding, view state"),
        ("Connection", "edge between two nodes", "source type, target type, data contract"),
        ("Canvas", "full dataflow", "mission, graph, selected region, all nodes, run state"),
        ("Data catalog", "dataset listing", "dataset metadata, provenance, permissions"),
        ("Document source", "PDF, web page, report", "extracted text, citations, source metadata"),
    ]
    for tgt, ex, ctx in hooks:
        ac.cell(row=r, column=2, value=tgt).font = font(10, True)
        ac.cell(row=r, column=2).border = BOX; ac.cell(row=r, column=2).alignment = WRAP
        ac.cell(row=r, column=3, value=ex).font = font(10); ac.cell(row=r, column=3).border = BOX; ac.cell(row=r, column=3).alignment = WRAP
        cc = ac.cell(row=r, column=4, value=ctx); cc.font = font(10); cc.border = BOX; cc.alignment = WRAP
        ac.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        for col in (5, 6):
            ac.cell(row=r, column=col).border = BOX
        ac.row_dimensions[r].height = est_height(ctx, 44)
        r += 1
    ac.freeze_panes = "A4"

    # ---- one sheet per flow --------------------------------------------
    for f in FLOWS:
        ws = wb.create_sheet(f"{f['id']} {f['title'].split(' — ')[0].split(' →')[0][:24]}")
        ws.sheet_properties.tabColor = f["accent"]
        setup_flow_columns(ws)
        # title
        ws.merge_cells("B1:D1")
        c = ws["B1"]; c.value = f"{f['id']} · {f['title']}"; c.font = font(15, True, WHITE); c.fill = fill(f["accent"])
        c.alignment = Alignment(vertical="center", indent=1); ws.row_dimensions[1].height = 30
        ws.merge_cells("B2:D2")
        c = ws["B2"]; c.value = f["intro"]; c.font = font(10, color=SUBTLE); c.alignment = WRAP
        ws.row_dimensions[2].height = est_height(f["intro"], 150, base=14)
        # image
        xl = XLImage(str(assets[f["id"]]))
        disp_w = 1080
        xl.width = disp_w
        xl.height = int(disp_w * 941 / 1672)
        ws.add_image(xl, "B4")
        img_rows = -(-xl.height // 20) + 1
        r = 4 + img_rows + 1
        # callouts table
        section_header(ws, r, "Callouts", f["accent"]); r += 1
        for i, cn in enumerate(["#", "UI element", "What it does"]):
            c = ws.cell(row=r, column=2 + i, value=cn); c.font = font(10, True, WHITE); c.fill = fill(SUBTLE)
            c.border = BOX; c.alignment = CENTER if i == 0 else WRAP_MID
        r += 1
        for i, (x, y, el, desc) in enumerate(f["callouts"], start=1):
            nb = ws.cell(row=r, column=2, value=i); nb.font = font(11, True, WHITE); nb.fill = fill(f["accent"])
            nb.alignment = CENTER; nb.border = BOX
            ce = ws.cell(row=r, column=3, value=el); ce.font = font(10, True); ce.alignment = WRAP; ce.border = BOX
            cd = ws.cell(row=r, column=4, value=desc); cd.font = font(10); cd.alignment = WRAP; cd.border = BOX
            ws.row_dimensions[r].height = est_height(desc, 118)
            r += 1
        r += 1
        section_header(ws, r, "Interactions & state changes", f["accent"]); r += 1
        r = bullets(ws, r, f["states"], f["accent"]); r += 1
        section_header(ws, r, "UX notes & refinements", f["accent"]); r += 1
        r = bullets(ws, r, f["notes"], f["accent"])

    # ---- Refinements ----------------------------------------------------
    rf = wb.create_sheet("Refinements")
    rf.sheet_properties.tabColor = ORANGE
    rf.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", (2.5, 6, 46, 34, 14)):
        rf.column_dimensions[col].width = w
    rf.merge_cells("B2:E2")
    h = rf["B2"]; h.value = "Refinements to align the screens with the concept"; h.font = font(16, True, WHITE); h.fill = fill(ORANGE)
    h.alignment = Alignment(vertical="center", indent=1); rf.row_dimensions[2].height = 32
    r = 4
    for i, cn in enumerate(["#", "Refinement", "Why", "Screens"]):
        c = rf.cell(row=r, column=2 + i, value=cn); c.font = font(10, True, WHITE); c.fill = fill(SUBTLE); c.border = BOX
        c.alignment = CENTER if i == 0 else WRAP_MID
    r += 1
    refs = [
        ("R1", "Selecting a node surfaces compatible agents in the catalog", "Discovery affordance; matches the concept", "01, 02"),
        ("R2", "Intent + context expressed as chat messages / quick replies", "Transparent, steerable automation — in chat, not a bespoke form", "03, 06, 07"),
        ("R3", "Order suggestions external-first, then catalog (in the chat card)", "Matches point 2", "03, 07"),
        ("R4", "Multi-select sources inside the chat suggestion card", "Matches point 6", "03, 07"),
        ("R5", "Dataset preview as an inline chat card (code + API key + sample)", "Matches point 7; removes the bespoke review modal", "08"),
        ("R6", "On confirm, create dataset nodes with an agent-provenance badge", "Matches points 3 & 6; legible provenance", "09"),
        ("R7", "Run history = chat history (transcript per node + agent)", "One model for point 5; no separate history panel", "06, 07, 09"),
        ("R8", "Keep confirmation explicit; show populated state; never silent", "Core review-before-apply invariant", "03, 08, 09"),
        ("R10", "Single, unified chat drawer for ALL attached agents", "Attachments are visually indistinguishable; differ only by agent type, config, and chat history", "03, 06, 07, 08, 09"),
        ("R11", "Session keyed by node id + agent type + project id; prev/next nav in the header", "Each attachment is a resumable session; navigate all agents in the dataflow", "03, 06, 07"),
        ("R12", "Remove bespoke refinement, source-review, and review-modal interfaces", "Reduce surface area; one chat pattern like Cursor / Claude", "03, 06, 07, 08"),
        ("R13", "Attached agents shown as a macOS-style dock: compact square icon-only tiles, hover magnification + name tooltip, running dots", "Familiar, compact, icon-first affordance consistent with the app style", "02, 03, 06"),
        ("R14", "Redefine Dataflow Builder as the MASTER ORCHESTRATOR: plans, spawns + coordinates specialized agents, evaluates, delivers the dataflow (execution-plan + spawned-agents chat cards)", "Catalog hosts all high-level agents; the builder orchestrates them rather than recommending isolated resources", "10"),
        ("R15", "Separate agent DEFINITION (LangChain, behind a swappable abstraction) from agent ASSIGNMENT (UI); prompts become classified reusable capabilities", "Keeps the UI framework-agnostic and extensible; see docs 09 & 10", "10"),
        ("R16", "Agents adopt the EXACT node/dataset install/publish UI: catalog cards with Install / Uninstall + v{N} pill + shared Publish → Published pill + Import agent footer; PUBLISH LIVES SOLELY IN THE DRAWER; palette rows mirror the dataset row exactly (icon + name + reference + category chip; no accent strip, no drag-handle icon, no per-row action); attachment is a palette drag", "Identical dynamics to how users install/publish nodes & datasets — same shared primitives; no palette publish control; no bespoke scope tiers; see docs 03 & 11", "01, 11"),
        ("R17", "Consolidate agent descriptions into manifest definitions (purpose, capabilities, targets, inputs, outputs, config, runtime, install_state/published/version) that drive catalog + palette + attachment + refinement", "Single source of truth; see doc 11", "01, 11"),
        ("R18", "Dataflow Builder proposes a reviewed Install in project for any specialist the project is missing — never an automatic or silent install; installed specialists then appear in the project palette", "Orchestration surfaces installs for review; agent installation is explicit, not automatic", "10, 11"),
        ("R19", "Dataset Finder suggests in TWO lanes from one discovery step: external sources and existing Data Catalog datasets, each with distinct per-row badges (source-type + confidence vs. format + install-state)", "Covers both discovery modes; the Finder stays focused on discovery + selection", "03, 07"),
        ("R20", "Selecting an EXTERNAL dataset hands off to a Node Builder agent that generates the complete fetch node (request code, params, auth/API key, parsing, error handling, output) — reviewed before it is added", "Separates discovery (Finder) from implementation (Node Builder); gives a real review-before-apply step for generated code", "07, 08"),
        ("R21", "Selecting a Data Catalog dataset reuses the EXISTING install flow (auto-install if not installed, project-scoped) and never generates external-fetch code; provenance in the palette distinguishes EXTERNAL (Node Builder) from IMPORTED (Data Catalog)", "No duplicated install/fetch logic; origins stay legible in the DATA palette", "07, 09"),
        ("R22", "Actions are SUGGESTED PROMPTS, not buttons: no bespoke per-agent controls (Create Node / Create Dataset / Add dataset node / card CTAs). Confirmations, options, follow-ups, and next steps are surfaced as suggested prompts — primary prefilled + editable in the chat input, alternatives as a chip row — that the user reviews, edits, and submits. One generic, reusable pattern for every agent", "Consistent, low-surface interaction; any agent presents actions the same way", "03, 06, 07, 08, 10"),
        ("R23", "Claude-like chat-feedback visual system on Curio tokens: subtle grouped surfaces + hairline borders (no heavy bars), leading accent dot for identity, raised inner panels, readable radio/checkbox option rows with soft accent selection states, gentle status/result tones, soft tokens/chips, polished spacing", "One calm, legible feedback language across every card and control; familiar agent-chat feel without a Curio-foreign look", "03, 06, 07, 08, 10"),
    ]
    for rid, what, why, scr in refs:
        rf.cell(row=r, column=2, value=rid).font = font(10, True); rf.cell(row=r, column=2).alignment = CENTER; rf.cell(row=r, column=2).border = BOX
        cw = rf.cell(row=r, column=3, value=what); cw.font = font(10, True); cw.alignment = WRAP; cw.border = BOX
        cy = rf.cell(row=r, column=4, value=why); cy.font = font(10); cy.alignment = WRAP; cy.border = BOX
        cs = rf.cell(row=r, column=5, value=scr); cs.font = font(10, color=SUBTLE); cs.alignment = WRAP; cs.border = BOX
        rf.row_dimensions[r].height = est_height(what, 44)
        r += 1
    r += 1
    section_header(rf, r, "Open questions", ORANGE)
    rf.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5); r += 1
    for it in [
        "Selection persistence: should selected sources survive a node-context change, or re-rank?",
        "Credential storage scope: per-dataflow vs per-account API keys.",
        "How many external sources to show before 'show more' (ranking cutoff)?",
        "Removing an agent after it populated a node — keep or revert the created datasets?",
    ]:
        cell = rf.cell(row=r, column=2, value="•  " + it); cell.font = font(10, color="3A3A42"); cell.alignment = WRAP
        rf.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        rf.row_dimensions[r].height = est_height(it, 150)
        r += 1
    rf.freeze_panes = "A4"

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX.relative_to(ROOT.parents[2]) if False else OUT_XLSX}")
    print(f"annotated {len(assets)} screens -> {ASSETS}")


if __name__ == "__main__":
    build()
